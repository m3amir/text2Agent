"""
SharePoint Tool using Microsoft Graph API
Provides folder listing and file search functionality
"""
import requests
import json
import asyncio
import io
from typing import Optional, Dict, List, Any

# Text extraction libraries
try:
    import docx2txt
    from PyPDF2 import PdfReader
    import openpyxl
    EXTRACTION_AVAILABLE = True
except ImportError:
    EXTRACTION_AVAILABLE = False

class SharepointToolkit:
    """SharePoint toolkit using Microsoft Graph API"""
    
    def __init__(self, credentials: Dict[str, str]):
        """Initialize with credentials"""
        self.tenant_id = credentials.get('tenant_id')
        self.client_id = credentials.get('client_id') 
        self.client_secret = credentials.get('client_secret')
        self.site_url = credentials.get('site_url')
        
        self.access_token = None
        self.site_id = None
        self.drives = {}
        
        # Validate required credentials
        if not all([self.tenant_id, self.client_id, self.client_secret, self.site_url]):
            raise ValueError("Missing required SharePoint credentials")
    
    async def _authenticate(self):
        """Get access token for Microsoft Graph"""
        if self.access_token:
            return True
            
        token_url = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        
        data = {
            'grant_type': 'client_credentials',
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'scope': 'https://graph.microsoft.com/.default'
        }
        
        try:
            response = requests.post(token_url, data=data)
            if response.status_code == 200:
                token_data = response.json()
                self.access_token = token_data.get('access_token')
                return True
            else:
                print(f"Authentication failed: {response.text}")
                return False
        except Exception as e:
            print(f"Authentication error: {e}")
            return False
    
    async def _get_site_info(self):
        """Get SharePoint site information"""
        if not await self._authenticate():
            return None
            
        if self.site_id:
            return self.site_id
        
        # Extract hostname and site path
        site_parts = self.site_url.replace('https://', '').split('/')
        hostname = site_parts[0]
        site_path = '/'.join(site_parts[1:])
        
        graph_url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'Content-Type': 'application/json'
        }
        
        try:
            response = requests.get(graph_url, headers=headers)
            if response.status_code == 200:
                site_data = response.json()
                self.site_id = site_data.get('id')
                return self.site_id
            else:
                print(f"Site lookup failed: {response.text}")
                return None
        except Exception as e:
            print(f"Site lookup error: {e}")
            return None
    
    async def _get_drives(self):
        """Get all drives (document libraries) in the site"""
        if not self.site_id:
            await self._get_site_info()
        
        if not self.site_id:
            return []
        
        drives_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives"
        
        headers = {
            'Authorization': f'Bearer {self.access_token}',
            'Content-Type': 'application/json'
        }
        
        try:
            response = requests.get(drives_url, headers=headers)
            if response.status_code == 200:
                drives_data = response.json()
                drives = drives_data.get('value', [])
                
                # Cache drives for easy lookup
                for drive in drives:
                    self.drives[drive.get('name')] = drive.get('id')
                
                return drives
            else:
                print(f"Drives lookup failed: {response.text}")
                return []
        except Exception as e:
            print(f"Drives lookup error: {e}")
            return []
    
    def sharepoint_list_folders(self, parent_folder: str = None, drive_name: str = "Documents") -> str:
        """
        List folders in SharePoint
        
        Args:
            parent_folder: Parent folder path (optional)
            drive_name: Name of the drive/library (default: Documents)
        
        Returns:
            JSON string with folder information
        """
        return asyncio.run(self._list_folders_async(parent_folder, drive_name))
    
    async def _list_folders_async(self, parent_folder: str = None, drive_name: str = "Documents") -> str:
        """Async implementation of folder listing"""
        try:
            # Get drives if not already loaded
            if not self.drives:
                await self._get_drives()
            
            # Get the specified drive
            drive_id = self.drives.get(drive_name)
            if not drive_id:
                return json.dumps({
                    "error": f"Drive '{drive_name}' not found",
                    "available_drives": list(self.drives.keys())
                })
            
            # Build URL for folder contents
            if parent_folder:
                folders_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{drive_id}/root:/{parent_folder}:/children"
            else:
                folders_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{drive_id}/root/children"
            
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(folders_url, headers=headers)
            
            if response.status_code == 200:
                items_data = response.json()
                items = items_data.get('value', [])
                
                # Filter only folders
                folders = []
                for item in items:
                    if 'folder' in item:
                        folders.append({
                            'name': item.get('name'),
                            'path': f"/{parent_folder}/{item.get('name')}" if parent_folder else f"/{item.get('name')}",
                            'created': item.get('createdDateTime'),
                            'modified': item.get('lastModifiedDateTime'),
                            'id': item.get('id'),
                            'childCount': item.get('folder', {}).get('childCount', 0)
                        })
                
                return json.dumps({
                    'success': True,
                    'drive': drive_name,
                    'parent_folder': parent_folder or 'root',
                    'folders': folders,
                    'count': len(folders)
                }, indent=2)
            else:
                return json.dumps({
                    'error': f"Failed to list folders: {response.status_code}",
                    'details': response.text
                })
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in list_folders: {str(e)}"
            })
    
    def sharepoint_search_files(self, query: str, drive_name: str = "Documents", file_type: str = None) -> str:
        """
        Search for files in SharePoint
        
        Args:
            query: Search query string
            drive_name: Name of the drive/library (default: Documents)
            file_type: File extension filter (e.g., 'pdf', 'docx')
        
        Returns:
            JSON string with search results
        """
        return asyncio.run(self._search_files_async(query, drive_name, file_type))
    
    async def _search_files_async(self, query: str, drive_name: str = "Documents", file_type: str = None) -> str:
        """Async implementation of file search"""
        try:
            # Get drives if not already loaded
            if not self.drives:
                await self._get_drives()
            
            # Get the specified drive
            drive_id = self.drives.get(drive_name)
            if not drive_id:
                return json.dumps({
                    "error": f"Drive '{drive_name}' not found",
                    "available_drives": list(self.drives.keys())
                })
            
            # Build search URL
            search_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{drive_id}/root/search(q='{query}')"
            
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(search_url, headers=headers)
            
            if response.status_code == 200:
                items_data = response.json()
                items = items_data.get('value', [])
                
                # Filter files (not folders) and by file type if specified
                files = []
                for item in items:
                    if 'file' in item:  # Only files, not folders
                        file_name = item.get('name', '')
                        
                        # Apply file type filter if specified
                        if file_type:
                            if not file_name.lower().endswith(f'.{file_type.lower()}'):
                                continue
                        
                        files.append({
                            'name': file_name,
                            'path': item.get('parentReference', {}).get('path', ''),
                            'size': item.get('size', 0),
                            'created': item.get('createdDateTime'),
                            'modified': item.get('lastModifiedDateTime'),
                            'id': item.get('id'),
                            'downloadUrl': item.get('@microsoft.graph.downloadUrl'),
                            'webUrl': item.get('webUrl'),
                            'mimeType': item.get('file', {}).get('mimeType')
                        })
                
                return json.dumps({
                    'success': True,
                    'query': query,
                    'drive': drive_name,
                    'file_type_filter': file_type,
                    'files': files,
                    'count': len(files)
                }, indent=2)
            else:
                return json.dumps({
                    'error': f"Search failed: {response.status_code}",
                    'details': response.text
                })
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in search_files: {str(e)}"
            })
    
    def sharepoint_get_drives(self) -> str:
        """
        Get all available drives/document libraries
        
        Returns:
            JSON string with drive information
        """
        return asyncio.run(self._get_drives_async())
    
    async def _get_drives_async(self) -> str:
        """Async implementation of get drives"""
        try:
            drives = await self._get_drives()
            
            drive_info = []
            for drive in drives:
                drive_info.append({
                    'name': drive.get('name'),
                    'id': drive.get('id'),
                    'description': drive.get('description'),
                    'driveType': drive.get('driveType'),
                    'webUrl': drive.get('webUrl')
                })
            
            return json.dumps({
                'success': True,
                'drives': drive_info,
                'count': len(drive_info)
            }, indent=2)
            
        except Exception as e:
            return json.dumps({
                'error': f"Exception in get_drives: {str(e)}"
            })
    
    def sharepoint_get_file_content(self, file_id: str, drive_name: str = "Documents") -> str:
        """
        Get file content/metadata
        
        Args:
            file_id: ID of the file
            drive_name: Name of the drive/library (default: Documents)
        
        Returns:
            JSON string with file information
        """
        return asyncio.run(self._get_file_content_async(file_id, drive_name))
    
    async def _get_file_content_async(self, file_id: str, drive_name: str = "Documents") -> str:
        """Async implementation of get file content"""
        try:
            # Get drives if not already loaded
            if not self.drives:
                await self._get_drives()
            
            # Get the specified drive
            drive_id = self.drives.get(drive_name)
            if not drive_id:
                return json.dumps({
                    "error": f"Drive '{drive_name}' not found",
                    "available_drives": list(self.drives.keys())
                })
            
            # Get file metadata
            file_url = f"https://graph.microsoft.com/v1.0/sites/{self.site_id}/drives/{drive_id}/items/{file_id}"
            
            headers = {
                'Authorization': f'Bearer {self.access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(file_url, headers=headers)
            
            if response.status_code == 200:
                file_data = response.json()
                
                file_info = {
                    'name': file_data.get('name'),
                    'size': file_data.get('size'),
                    'created': file_data.get('createdDateTime'),
                    'modified': file_data.get('lastModifiedDateTime'),
                    'id': file_data.get('id'),
                    'downloadUrl': file_data.get('@microsoft.graph.downloadUrl'),
                    'webUrl': file_data.get('webUrl'),
                    'mimeType': file_data.get('file', {}).get('mimeType'),
                    'path': file_data.get('parentReference', {}).get('path', '')
                }
                
                return json.dumps({
                    'success': True,
                    'file': file_info
                }, indent=2)
            else:
                return json.dumps({
                    'error': f"File lookup failed: {response.status_code}",
                    'details': response.text
                })
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in get_file_content: {str(e)}"
            })
    
    def sharepoint_get_document_by_name(self, document_name: str, drive_name: str = "Documents", exact_match: bool = True) -> str:
        """
        Get a document by its name
        
        Args:
            document_name: Name of the document to search for
            drive_name: Name of the drive/library (default: Documents)
            exact_match: If True, search for exact name match; if False, search for partial match
        
        Returns:
            JSON string with document information
        """
        return asyncio.run(self._get_document_by_name_async(document_name, drive_name, exact_match))
    
    async def _get_document_by_name_async(self, document_name: str, drive_name: str = "Documents", exact_match: bool = True) -> str:
        """Async implementation of get document by name"""
        try:
            # Use the existing search functionality
            search_result = await self._search_files_async(document_name, drive_name)
            search_data = json.loads(search_result)
            
            if not search_data.get('success'):
                return search_result
            
            files = search_data.get('files', [])
            
            if exact_match:
                # Filter for exact name match
                exact_matches = [f for f in files if f.get('name', '').lower() == document_name.lower()]
                
                if len(exact_matches) == 1:
                    return json.dumps({
                        'success': True,
                        'document_name': document_name,
                        'exact_match': True,
                        'file': exact_matches[0]
                    }, indent=2)
                elif len(exact_matches) > 1:
                    return json.dumps({
                        'success': True,
                        'document_name': document_name,
                        'exact_match': True,
                        'multiple_matches': True,
                        'files': exact_matches,
                        'count': len(exact_matches)
                    }, indent=2)
                else:
                    return json.dumps({
                        'success': True,
                        'document_name': document_name,
                        'exact_match': True,
                        'found': False,
                        'message': f"No exact match found for '{document_name}'",
                        'suggestion': f"Try using exact_match=False for partial matching"
                    }, indent=2)
            else:
                # Return all partial matches
                return json.dumps({
                    'success': True,
                    'document_name': document_name,
                    'exact_match': False,
                    'files': files,
                    'count': len(files)
                }, indent=2)
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in get_document_by_name: {str(e)}"
            })
    
    def sharepoint_get_document_by_name_and_type(self, document_name: str, file_type: str, drive_name: str = "Documents") -> str:
        """
        Get a document by its name and file type
        
        Args:
            document_name: Name pattern of the document to search for
            file_type: File extension (e.g., 'pdf', 'docx')
            drive_name: Name of the drive/library (default: Documents)
        
        Returns:
            JSON string with document information
        """
        return asyncio.run(self._get_document_by_name_and_type_async(document_name, file_type, drive_name))
    
    async def _get_document_by_name_and_type_async(self, document_name: str, file_type: str, drive_name: str = "Documents") -> str:
        """Async implementation of get document by name and type"""
        try:
            # Use the existing search functionality with file type filter
            search_result = await self._search_files_async(document_name, drive_name, file_type)
            search_data = json.loads(search_result)
            
            if not search_data.get('success'):
                return search_result
            
            files = search_data.get('files', [])
            
            # Look for exact name matches within the filtered results
            exact_matches = []
            partial_matches = []
            
            for file in files:
                file_name = file.get('name', '')
                if file_name.lower() == f"{document_name.lower()}.{file_type.lower()}":
                    exact_matches.append(file)
                else:
                    partial_matches.append(file)
            
            result = {
                'success': True,
                'document_name': document_name,
                'file_type': file_type,
                'drive': drive_name
            }
            
            if exact_matches:
                if len(exact_matches) == 1:
                    result['exact_match_found'] = True
                    result['file'] = exact_matches[0]
                else:
                    result['multiple_exact_matches'] = True
                    result['exact_matches'] = exact_matches
                    result['exact_count'] = len(exact_matches)
            
            if partial_matches:
                result['partial_matches'] = partial_matches
                result['partial_count'] = len(partial_matches)
            
            if not exact_matches and not partial_matches:
                result['found'] = False
                result['message'] = f"No documents found matching '{document_name}' with type '{file_type}'"
            
            return json.dumps(result, indent=2)
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in get_document_by_name_and_type: {str(e)}"
            })
    
    def sharepoint_download_and_extract_text(self, file_id: str, drive_name: str = "Documents") -> str:
        """
        Download a document and extract its text content
        
        Args:
            file_id: ID of the file to download and extract text from
            drive_name: Name of the drive/library (default: Documents)
        
        Returns:
            JSON string with extracted text content
        """
        return asyncio.run(self._download_and_extract_text_async(file_id, drive_name))
    
    async def _download_and_extract_text_async(self, file_id: str, drive_name: str = "Documents") -> str:
        """Async implementation of download and extract text"""
        try:
            if not EXTRACTION_AVAILABLE:
                return json.dumps({
                    'error': 'Text extraction libraries not available',
                    'message': 'Install required packages: pip install python-docx PyPDF2 openpyxl docx2txt',
                    'success': False
                })
            
            # First get file metadata
            file_info_result = await self._get_file_content_async(file_id, drive_name)
            file_info_data = json.loads(file_info_result)
            
            if not file_info_data.get('success'):
                return file_info_result
            
            file_info = file_info_data['file']
            download_url = file_info.get('downloadUrl')
            
            if not download_url:
                return json.dumps({
                    'error': 'No download URL available for this file',
                    'success': False
                })
            
            # Download the file
            headers = {
                'Authorization': f'Bearer {self.access_token}'
            }
            
            response = requests.get(download_url, headers=headers)
            
            if response.status_code != 200:
                return json.dumps({
                    'error': f"Failed to download file: {response.status_code}",
                    'details': response.text,
                    'success': False
                })
            
            # Extract text based on file type
            file_name = file_info.get('name', '').lower()
            mime_type = file_info.get('mimeType', '')
            
            extracted_text = ""
            extraction_method = ""
            
            try:
                if file_name.endswith('.docx') or 'wordprocessingml' in mime_type:
                    # Extract from Word document
                    file_stream = io.BytesIO(response.content)
                    extracted_text = docx2txt.process(file_stream)
                    extraction_method = "docx2txt"
                    
                elif file_name.endswith('.pdf') or mime_type == 'application/pdf':
                    # Extract from PDF
                    file_stream = io.BytesIO(response.content)
                    pdf_reader = PdfReader(file_stream)
                    extracted_text = ""
                    for page_num, page in enumerate(pdf_reader.pages):
                        extracted_text += f"\n--- Page {page_num + 1} ---\n"
                        extracted_text += page.extract_text()
                    extraction_method = "PyPDF2"
                    
                elif file_name.endswith(('.xlsx', '.xls')) or 'spreadsheetml' in mime_type:
                    # Extract from Excel
                    file_stream = io.BytesIO(response.content)
                    workbook = openpyxl.load_workbook(file_stream, data_only=True)
                    extracted_text = ""
                    
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]
                        extracted_text += f"\n--- Sheet: {sheet_name} ---\n"
                        
                        for row in sheet.iter_rows(values_only=True):
                            row_text = []
                            for cell in row:
                                if cell is not None:
                                    row_text.append(str(cell))
                            if row_text:
                                extracted_text += " | ".join(row_text) + "\n"
                    
                    extraction_method = "openpyxl"
                    
                elif file_name.endswith('.txt') or mime_type == 'text/plain':
                    # Extract from text file
                    extracted_text = response.content.decode('utf-8')
                    extraction_method = "plain text"
                    
                else:
                    return json.dumps({
                        'error': f"Unsupported file type for text extraction: {file_name}",
                        'supported_types': ['.docx', '.pdf', '.xlsx', '.xls', '.txt'],
                        'file_mime_type': mime_type,
                        'success': False
                    })
                
                # Clean up the extracted text
                extracted_text = extracted_text.strip()
                
                return json.dumps({
                    'success': True,
                    'file_info': {
                        'name': file_info.get('name'),
                        'size': file_info.get('size'),
                        'mime_type': mime_type,
                        'id': file_id
                    },
                    'extraction': {
                        'method': extraction_method,
                        'text_length': len(extracted_text),
                        'text': extracted_text
                    },
                    'preview': extracted_text[:500] + "..." if len(extracted_text) > 500 else extracted_text
                }, indent=2)
                
            except Exception as extraction_error:
                return json.dumps({
                    'error': f"Text extraction failed: {str(extraction_error)}",
                    'extraction_method': extraction_method,
                    'file_type': file_name,
                    'success': False
                })
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in download_and_extract_text: {str(e)}",
                'success': False
            })
    
    def sharepoint_extract_text_by_name(self, document_name: str, drive_name: str = "Documents", exact_match: bool = True) -> str:
        """
        Find a document by name and extract its text content
        
        Args:
            document_name: Name of the document to find and extract text from
            drive_name: Name of the drive/library (default: Documents)
            exact_match: If True, search for exact name match; if False, search for partial match
        
        Returns:
            JSON string with extracted text content
        """
        return asyncio.run(self._extract_text_by_name_async(document_name, drive_name, exact_match))
    
    async def _extract_text_by_name_async(self, document_name: str, drive_name: str = "Documents", exact_match: bool = True) -> str:
        """Async implementation of extract text by name"""
        try:
            # First find the document
            search_result = await self._get_document_by_name_async(document_name, drive_name, exact_match)
            search_data = json.loads(search_result)
            
            if not search_data.get('success'):
                return search_result
            
            # Check if we found exactly one file
            if 'file' in search_data:
                # Single file found
                file_id = search_data['file']['id']
                return await self._download_and_extract_text_async(file_id, drive_name)
            elif 'files' in search_data and len(search_data['files']) == 1:
                # Single file from multiple results
                file_id = search_data['files'][0]['id']
                return await self._download_and_extract_text_async(file_id, drive_name)
            elif 'files' in search_data and len(search_data['files']) > 1:
                # Multiple files found
                return json.dumps({
                    'error': 'Multiple files found. Please be more specific or use exact file ID.',
                    'found_files': [{'name': f['name'], 'id': f['id']} for f in search_data['files']],
                    'suggestion': 'Use sharepoint_download_and_extract_text() with a specific file ID',
                    'success': False
                })
            else:
                # No files found
                return json.dumps({
                    'error': f"No document found with name: {document_name}",
                    'success': False
                })
                
        except Exception as e:
            return json.dumps({
                'error': f"Exception in extract_text_by_name: {str(e)}",
                'success': False
            }) 